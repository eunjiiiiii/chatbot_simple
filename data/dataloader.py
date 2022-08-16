import openpyxl
import random
from openpyxl import Workbook, load_workbook
# from kobert_transformers import get_tokenizer
from kogpt2_transformers import get_kogpt2_tokenizer


def wellness_question_data():
  """

  웰니스 데이터에서 (챗봇)질문만 뽑아서 wellness_dialog_question.txt 데이터 생성

  """
  root_path = "C:/Users/R301-6/Desktop/chatbot-master/demo/data/raw"
  wellness_file = root_path + "/wellness_dialog_dataset.xlsx"
  wellness_q_output = root_path + "/wellness_dialog_question.txt"

  f = open(wellness_q_output, 'w')

  wb = load_workbook(filename=wellness_file)

  ws = wb[wb.sheetnames[0]]
  # print(sheet)
  for row in ws.iter_rows(min_row=2):
    f.write(row[0].value + "    " + row[1].value + "\n")

  f.close()

def wellness_answer_data():
  """

  웰니스 데이터에서 (챗봇)답변만 뽑아서 wellness_dialog_answer.txt 데이터 생성

  """
  root_path = "C:/Users/R301-6/Desktop/chatbot-master/demo/data/raw"
  wellness_file = root_path + "/wellness_dialog_dataset.xlsx"
  wellness_a_output = root_path + "/wellness_dialog_answer.txt"

  f = open(wellness_a_output, 'w')
  wb = load_workbook(filename=wellness_file)
  ws = wb[wb.sheetnames[0]]

  for row in ws.iter_rows(min_row=2):
    if row[2].value == None:
      continue
    else:
      f.write(row[0].value + "    " + row[2].value + "\n")
  f.close()

def xlsx_to_text():
  """

  wellness_dialog_dataset.xlsx을 chatbot_wellness_data.txt로 바꾸기

  """
  root_path = "C:/Users/R301-6/Desktop/chatbot-master/demo/data/raw"
  wellness_file = root_path + "/wellness_dialog_dataset.xlsx"
  wellness_t_output = root_path + "/chatbot_wellness_data.txt"

  f = open(wellness_t_output, 'w')
  wb = load_workbook(filename=wellness_file)
  ws = wb[wb.sheetnames[0]]

  for row in ws.iter_rows(min_row=2):
    if row[2].value == None:
      continue
    else:
      f.write(row[0].value + "    " + row[1].value + "    " + row[2].value + "\n")
  f.close()

def category_data():
  """
  카테고리 클래스 데이터(chatbot_wellness_category.txt) : 카테고리 클래스 359개

  ex) 감정/감정조절이상 0
      감정/감정조잘이상/화 1
      감정/걱정 2
  """
  root_path = "C:/Users/R301-6/Desktop/chatbot-master/demo/data/raw"
  data_path = root_path + "/chatbot_wellness_data.txt"
  c_output = root_path + "/chatbot_wellness_category.txt"

  i_f = open(data_path, 'r')
  o_f = open(c_output, 'w')

  category_count = 0
  flag = True

  cate_dict = []
  i_lines = i_f.readlines()
  for i, data in enumerate(i_lines):
    tmp = data.split('    ')
    a = tmp[1][:-1]
    q = tmp[0]
    if a not in cate_dict:
      cate_dict.append(a)
      o_f.write(a.strip() + "    " + str(category_count) + "\n")
      category_count += 1
  o_f.close()
  i_f.close()

def wellness_text_classification_data():
  """
  카테고리 클래스와 질의 데이터 생성(wellness_dialog_for_text_classification.txt)

  ex) 감정/감정조절이상    그럴 때는 밥은 잘 먹었는지, 잠은 잘 잤는지 체크해보는 것도 좋아요.
  """

  root_path = "C:/Users/R301-6/Desktop/chatbot-master/demo/data/raw"
  wellness_category_file = root_path + "/wellness_dialog_category.txt"
  wellness_question_file = root_path + "/wellness_dialog_question.txt"
  wellness_text_classification_file = root_path + "/wellness_dialog_for_text_classification.txt"

  cate_file = open(wellness_category_file, 'r')
  ques_file = open(wellness_question_file, 'r')
  text_classfi_file = open(wellness_text_classification_file, 'w')

  category_lines = cate_file.readlines()
  cate_dict = {}
  for line_num, line_data in enumerate(category_lines):
    data = line_data.split('    ')
    cate_dict[data[0]] = data[1][:-1]
  print(cate_dict)

  ques_lines = ques_file.readlines()
  ques_dict = {}
  for line_num, line_data in enumerate(ques_lines):
    data = line_data.split('    ')
    # print(data[1]+ "    " + cate_dict[data[0]])
    text_classfi_file.write(data[1][:-1] + "    " + cate_dict[data[0]] + "\n")

  cate_file.close()
  ques_file.close()
  text_classfi_file.close()



def wellness_dialog_for_autoregressive():
  """
  카테고리/질문/답변 데이터에서 질문과 답변 쌍으로 구성
  (wellness_dialog_for_autoregressive.txt)
  """
  root_path = "C:/Users/R301-6/Desktop/chatbot-master/demo/data/raw"
  wellness_file = root_path + "/wellness_dialog_dataset.xlsx"
  wellness_answer_file = root_path + "/wellness_dialog_answer.txt"
  wellness_question_file = root_path + "/wellness_dialog_question.txt"
  wellness_autoregressive_file = root_path + "/wellness_dialog_for_autoregressive.txt"


  answ_file = open(wellness_answer_file, 'r')
  ques_file = open(wellness_question_file, 'r')
  autoregressive_file = open(wellness_autoregressive_file, 'w')

  answ_lines = answ_file.readlines()
  ques_lines = ques_file.readlines()
  ques_dict = {}
  for line_num, line_data in enumerate(ques_lines):
    ques_data = line_data.split('    ')
    for ans_line_num, ans_line_data in enumerate(answ_lines):
      ans_data = ans_line_data.split('    ')
      if ques_data[0] == ans_data[0]:
        autoregressive_file.write(ques_data[1][:-1] + "    " + ans_data[1])
      else:
        continue

  answ_file.close()
  ques_file.close()
  autoregressive_file.close()


def seperate_wellness_data():
  """
  wellness_dialog_for_autoregressive.txt 에서

  train set, test set split

  (wellness_dialog_for_autoregressive_train.txt,
   wellness_dialog_for_autoregressive_test.txt)
  """

  # wellness_autoregressive_file = root_path+"/wellness_dialog_for_autoregressive.txt"
  # wellness_text_classification_file = root_path + "/wellness_dialog_for_text_classification.txt"
  file_path = root_path + "/wellness_dialog_for_autoregressive.txt"
  train_file_path = root_path + "/wellness_dialog_for_autoregressive_train.txt"
  test_file_path = root_path + "/wellness_dialog_for_autoregressive_test.txt"

  sperated_file = open(file_path, 'r')
  train_file = open(train_file_path, 'w')
  test_file = open(test_file_path, 'w')

  sperated_file_lines = sperated_file.readlines()
  ques_dict = {}
  for line_num, line_data in enumerate(sperated_file_lines):
    rand_num = random.randint(0, 10)
    if rand_num < 10:
      train_file.write(line_data)
    else:
      test_file.write(line_data)

  sperated_file.close()
  train_file.close()
  test_file.close()


if __name__ == "__main__":

  wellness_question_data()
  wellness_answer_data()
  xlsx_to_text()
  category_data()
  wellness_dialog_for_autoregressive()
  #seperate_wellness_data()

  root_path = "C:/Users/R301-6/Desktop/chatbot-master/demo/data/raw"
  file_path = root_path + "/chatbot_wellness_data.txt"
  o_path = root_path + "/chatbot_wellness_data_for_autoregressive.txt"

  i_file = open(file_path, 'r')
  o_file = open(o_path, 'w')

  i_lines = i_file.readlines()
  for i, data in enumerate(i_lines):
    tmp = data.split('    ')
    question = tmp[0]
    answer = tmp[1][:-1]
    o_file.write("<s>" + question + "</s><s>" + answer+ "</s>\n")